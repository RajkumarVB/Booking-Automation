from openpyxl import Workbook,load_workbook
import pandas as pd
import json
import os
from datetime import datetime
from flask import Flask, jsonify, request, Response

data = []  
app = Flask(__name__)

def tatkalsheetFinder(panel_name):
    switcher = {
        "tatkal": "Tatkal_Details.xlsx",
        "tatkal_food": "Tatkal_Details.xlsx"
    }
    return switcher.get(panel_name,"nothing")

def tirupatisheetFinder(panel_name):
    switcher = {
        "seva": "Seva_details.xlsx",
        "300": "300rs_ticket_details.xlsx",
        "accomodation": "Accomodation_details.xlsx",
        "seva_login": "Login_details.xlsx",
        "credit_card":"cc_details.xlsx"
    }
    return switcher.get(panel_name,"nothing")

@app.route('/<panel_name>', methods = ['GET'])
def getTatkalDetails(panel_name):
    if tatkalsheetFinder(panel_name) == "nothing":
        print("Invalid panel name")
        return FileNotFoundError
    workbook = load_workbook(tatkalsheetFinder(panel_name))
    TatkalDict = {}
    sheets_list = workbook.sheetnames
    for sheet in sheets_list:
        df = pd.read_excel(tatkalsheetFinder(panel_name),sheet_name=sheet,usecols=['Name','Age'])
        df = df.dropna(axis='rows',subset=['Name'])
        if(df.empty == 1):
            continue
        TatkalDict[sheet] = df.to_dict(orient="records")
    return jsonify(TatkalDict)

@app.route('/tirupati/<panel_name>', methods = ['GET'])
def getTirupatiDetails(panel_name):
    match panel_name:
        case "seva_login":
            df = pd.read_excel(tirupatisheetFinder(panel_name),sheet_name="Sheet1",usecols=['Login_id'])
            df = df.dropna(axis='rows',subset=['Login_id'])
            df = df['Login_id'].str.strip()
            result = df.to_json(orient="index")
            parsed = json.loads(result)
            return jsonify(parsed)
        case "credit_card":
            df = pd.read_excel(tirupatisheetFinder(panel_name),sheet_name="Sheet1",usecols=['cc_no'])
            df = df.dropna(axis='rows',subset=['cc_no'])
            df['cc_no'] = df['cc_no'].astype(str).apply(lambda x: x.replace('.0',''))
            df = df['cc_no'].str.strip()
            result = df.to_json(orient="index")
            parsed = json.loads(result)
            return jsonify(parsed)
        case other:
            workbook = load_workbook(tirupatisheetFinder(panel_name))
            TptyDict = {}
            sheets_list = workbook.sheetnames
            print(type(sheets_list))
            for sheet in sheets_list:
                df = pd.read_excel(tirupatisheetFinder(panel_name),sheet_name=sheet,usecols=['Name','Proof','Proof_no'])
                df = df.dropna(axis='rows',subset=['Name'])
                if(df.empty == 1):
                    continue
                # sheet_no = str(sheet)
                TptyDict[sheet] = df.to_dict(orient="records")
            return jsonify(TptyDict)

@app.route('/execute/<panel_name>/<int:sheet_no>', methods = ['GET'])
def execute_script(panel_name,sheet_no):
    match panel_name:
        case "tatkal":
            os.system(f"python taktal_automate.py {sheet_no}")
        case "seva":
            os.system(f"python seva_automate.py {sheet_no}")
        case "300":
            os.system(f"python 300_tpty.py {sheet_no}")
        case "accomodation":
            os.system(f"python accomodation_automate.py {sheet_no}")
        case "seva_login":
            os.system(f"python login_automate.py {sheet_no}")
        case "credit_card":
            os.system(f"python cc_automate.py {sheet_no}") 
        case "tatkal_food":
            os.system(f"python tatkal_food.py {sheet_no}")
        case other:
            print("error: panel not found")
    return history(panel_name,sheet_no)
    #return jsonify("status_code: 200")

def history(panel_name,sheet_no):
    global data
    if len(data)> 10:
       data.pop(1)   
    currentData = {}
    currentData["panel_name"] = panel_name
    currentData["sheet_no"] = sheet_no
    currentData["timestamp"] = str(datetime.now())
    data.append(currentData)
    #print(f"Current Data = {currentData}")
    #print(f"Data = {data}")
    data = sorted(data, key=lambda x: x["timestamp"], reverse=True)
    return jsonify(data)  
                      
if __name__ == "__main__":
    app.run(debug=True)



